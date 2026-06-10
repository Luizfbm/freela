from pathlib import Path
import json
import re

from openpyxl import load_workbook


ROOT = Path("/Users/luiz_fbm/Documents/programacao/freela")
OUTPUT_DIR = ROOT / "outputs" / "leads_saude_consolidado_20260610"
OUTPUT_JSON = OUTPUT_DIR / "merged_leads.json"

SOURCES = [
    {
        "origin": "Vila Velha/ES",
        "path": ROOT
        / "outputs"
        / "vila_velha_saude_leads_20260609"
        / "leads_saude_vila_velha_es_2026-06-09.xlsx",
        "kind": "vila_velha",
    },
    {
        "origin": "Vitoria/ES",
        "path": ROOT
        / "outputs"
        / "vitoria_saude_leads_20260605"
        / "leads_saude_vitoria_es_2026-06-05.xlsx",
        "kind": "vitoria",
    },
]


def clean(value):
    if value is None:
        return ""
    return str(value).strip()


def normalize_key(value):
    value = clean(value).lower()
    value = re.sub(r"\s+", " ", value)
    return value


def segment_from_category(category):
    text = normalize_key(category)
    if "odonto" in text or "dent" in text or "ortodont" in text:
        return "Odontologia"
    if "laboratorio" in text or "analise" in text or "citologia" in text:
        return "Laboratorio"
    if "pilates" in text:
        return "Pilates"
    if "fisio" in text:
        return "Fisioterapia"
    return "Saude"


def read_rows(source):
    wb = load_workbook(source["path"], data_only=True)
    ws = wb["Leads"]
    headers = [clean(cell.value) for cell in ws[1]]
    records = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not any(row):
            continue
        raw = {headers[index]: row[index] for index in range(min(len(headers), len(row)))}
        if source["kind"] == "vila_velha":
            website_url = clean(raw.get("website_url"))
            social_urls = clean(raw.get("social_urls"))
            website_or_social = "; ".join(
                item
                for item in [website_url if website_url != "Not found" else "", social_urls if social_urls != "Not found" else ""]
                if item
            ) or "Not found"
            category = clean(raw.get("category"))
            records.append(
                {
                    "origin": source["origin"],
                    "score": clean(raw.get("score")),
                    "business": clean(raw.get("business")),
                    "segment": segment_from_category(category),
                    "category": category,
                    "area": clean(raw.get("area")),
                    "distance": f"{clean(raw.get('distance_km'))} km" if clean(raw.get("distance_km")) else "",
                    "website_status": clean(raw.get("website_status")),
                    "website_or_social": website_or_social,
                    "phone": clean(raw.get("phone")),
                    "source_urls": clean(raw.get("source_urls")),
                    "why_prospect": clean(raw.get("why_prospect")),
                    "confidence": clean(raw.get("confidence")),
                    "notes": clean(raw.get("notes")),
                    "next_action": "Validar dados finais e enviar demo personalizada.",
                }
            )
        else:
            records.append(
                {
                    "origin": source["origin"],
                    "score": clean(raw.get("Score")),
                    "business": clean(raw.get("Negócio")),
                    "segment": clean(raw.get("Segmento")),
                    "category": clean(raw.get("Categoria")),
                    "area": clean(raw.get("Área")),
                    "distance": clean(raw.get("Distância")),
                    "website_status": clean(raw.get("Status do site")),
                    "website_or_social": clean(raw.get("Website/Social")),
                    "phone": clean(raw.get("Telefone")),
                    "source_urls": clean(raw.get("Fonte principal")),
                    "why_prospect": clean(raw.get("Por que é prospect")),
                    "confidence": clean(raw.get("Confiança")),
                    "notes": clean(raw.get("Notas")),
                    "next_action": clean(raw.get("Próxima ação")),
                }
            )
    return records


def main():
    merged = []
    seen = set()
    for source in SOURCES:
        for record in read_rows(source):
            key = (normalize_key(record["business"]), normalize_key(record["phone"]))
            if key in seen:
                continue
            seen.add(key)
            merged.append(record)

    payload = {
        "sources": [
            {"origin": source["origin"], "path": str(source["path"])} for source in SOURCES
        ],
        "rows": merged,
    }
    OUTPUT_JSON.write_text(json.dumps(payload, ensure_ascii=False, indent=2) + "\n", encoding="utf-8")
    print(json.dumps({"rows": len(merged), "output": str(OUTPUT_JSON)}, ensure_ascii=False))


if __name__ == "__main__":
    main()
